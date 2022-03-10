# Project 0 to open excel file and read data from it

import openpyxl


def open_file():
    file_path = input("enter the file path: ")
    file = openpyxl.load_workbook(file_path)
    global sheet
    sheet = file.active
    global lastrow
    lastrow = sheet.max_row


def displaying_content(lastrow, sheet):
    for r in range(1, lastrow + 1):
        state = chr(65)  # this should read the date from the first column
        population = chr(66)  # this should read the date from the second column
        index_state = state + str(r)
        population_state = population + str(r)
        print(sheet[index_state].value, " ",
              sheet[population_state].value)  # this should print the state with it's population


def appending_values(lastrow):
    global populations, states
    populations = []
    states = []
    for k in range(1, lastrow + 1):
        populations.append(sheet.cell(row=k, column=2).value)
        states.append(sheet.cell(row=k, column=1).value)


def find_max_population(lastrow, populations, states):  # this should find the country with the greatest population
    # those are filled later to be used with the max() function

    max_population = max(populations)
    index_max_state = populations.index(max_population)

    print("The state with max population is:", states[index_max_state])


def find_min_population(lastrow, populations,states):  
    # this should work like the find_max population function but with min()

    min_population = min(populations)
    index_min_state = populations.index(
        min_population)  # this should get the index of the state with the highest population

    print("The state with min population is:", states[index_min_state])


def choices():
    open_file()
    print("Enter 1 to display population of each state: ")
    print("Enter 2 to display the state with the highest and lowest population: ")
    print("Enter 3 to exit")
    choice = int(input())

    if choice == 1:
        displaying_content(lastrow, sheet)
    elif choice == 2:
        appending_values(lastrow)
        find_max_population(lastrow, populations, states)
        find_min_population(lastrow, populations, states)
    else:
        quit()


try:
    choices()
except FileNotFoundError:
    print("Country doesn't exist")
