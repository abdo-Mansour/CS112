# This is Group task 0, Assignment 1
# CS112
# Created by Mansour(20210057) on 28/2/2022

import openpyxl

file_location = input("Please enter the path of the file: ")


def openinig_file():  # This should execute after checking if the file exists or not
    file = openpyxl.load_workbook(file_location)

    content = file.active

    max_rows = content.max_row


def display_content(max_rows):  # this should print all the value of the states and their population
    for i in range(1, max_rows + 1):
        state = content.cell(row=i, column=1)
        population = content.cell(row=i, column=2)
        print(state.value, ' ', population.value)


def find_max_population(max_rows):  # this should find the country with the greatest population
    # those are filled later to be used with the max() function
    populations = []
    states = []

    for k in range(1, max_rows + 1):
        populations.append(content.cell(row=k, column=2).value)
        states.append(content.cell(row=k, column=1).value)

    max_population = max(populations)
    index_max_state = populations.index(max_population)

    print("The state with max population is:", states[index_max_state])


def find_min_population(max_rows):  # this should work like the find_max population function but with min()
    populations = []
    states = []

    for k in range(1, max_rows + 1):
        populations.append(content.cell(row=k, column=2).value)
        states.append(content.cell(row=k, column=1).value)

    min_population = min(populations)
    index_min_state = populations.index(
        min_population)  # this should get the index of the state with the highest population

    print("The state with min population is:", states[index_min_state])


def whole_game():
    openinig_file()
    display_content(max_rows)
    find_max_population(max_rows)
    find_min_population(max_rows)


try:
    whole_game()
except FileNotFoundError:
    print("Country doesn't exist")
